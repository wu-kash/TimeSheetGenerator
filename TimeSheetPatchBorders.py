# -*- coding: utf-8 -*-
"""
Created on Thu Nov 22 17:51:57 2018

@author: Lukasz
"""
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill
import re

indexLetter = {
               'A': 1, 'B': 2, 'C': 3, 'D': 4,
               'E': 5, 'F': 6, 'G': 7, 'H': 8,
               'I': 9, 'J': 10, 'K': 11, 'L': 12,
               'M': 13, 'N': 14, 'O': 15, 'P': 16,
               'Q': 17, 'R': 18, 'S': 19, 'T': 20,
               'U': 21, 'V': 22, 'W': 23, 'X': 24,
               'Y': 25, 'Z': 26
               }


def fillCellColour( cell1, cell2, colour, sheet):
    
    if colour == 'whiteFill':
        colourFill = PatternFill(start_color='FFFFFFFF',
                   end_color='FFFFFFFF',
                   fill_type='solid')
    else:
        colourFill = PatternFill(start_color='00D3D3D3',
                   end_color='00D3D3D3',
                   fill_type='solid')

    
    columnStart = int(indexLetter[re.split('\d+', cell1)[0]])
    rowStart = int(re.split('(\d+)', cell1)[1])
    columnEnd = int(indexLetter[re.split('\d+', cell2)[0]])
    
    for i in range(columnStart, columnEnd + 1):
        for letter, index in indexLetter.items():
            if index == i:
                sheet[str(letter) + str(rowStart)].fill = colourFill
    

def patchCellBorder( cell1, cell2, borderStyle, sheet):
    
    'Letter is the column, number is the row'
    columnStart = int(indexLetter[re.split('\d+', cell1)[0]])
    rowStart = int(re.split('(\d+)', cell1)[1])
    columnEnd = int(indexLetter[re.split('\d+', cell2)[0]])
    rowEnd = int(re.split('(\d+)', cell2)[1])
    
    if (columnStart is columnEnd):
        for i in range(rowStart, rowEnd + 1 ):
            sheet.cell(row = i, column = columnStart).border = borderStyle
    else:
        for i in range(columnStart, columnEnd + 1 ):
            sheet.cell(row = rowStart, column = i).border = borderStyle
                

def patchSheetBorder(sheet):
    companyBorder = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    nameBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='dashed'))
    IDBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='dotted'))
    surnameBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='dotted'), 
                         bottom=Side(style='medium'))
    contactBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='dotted'), 
                         bottom=Side(style='medium'))
    dateBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))
    topBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))
    lineBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='dashed'), 
                         bottom=Side(style='dashed'))
    bottomBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='dashed'), 
                         bottom=Side(style='medium'))
    totalBorder = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))
    
    
    patchCellBorder('C2', 'R2', companyBorder, sheet) 
    patchCellBorder('C3', 'J3', nameBorder, sheet)
    patchCellBorder('K3', 'R3', IDBorder, sheet)
    patchCellBorder('C4', 'J4', surnameBorder, sheet)
    patchCellBorder('K4', 'R4', contactBorder, sheet)
    patchCellBorder('C5', 'R5', dateBorder, sheet)
    patchCellBorder('B7', 'S7', topBorder, sheet)
    patchCellBorder('B8', 'S8', topBorder, sheet)
    for i in range(9, 40):
        patchCellBorder( str('B' + str(i)), str('S' + str(i)), lineBorder, sheet)
    patchCellBorder('B39', 'S39', bottomBorder, sheet)
    patchCellBorder('P40', 'S40', totalBorder, sheet)